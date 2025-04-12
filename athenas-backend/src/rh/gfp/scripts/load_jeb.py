import codecs
import datetime

from contrib.middleware import set_current_user
from rh.gfp.models import CorrectionFactor
from openpyxl import load_workbook

set_current_user("athenas")

file_path = "/home/raysonsilva/Downloads/JEBR0622N.csv"

idx_months = {
    "JAN": 1,
    "FEV": 2,
    "MAR": 3,
    "ABR": 4,
    "MAI": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SET": 9,
    "OUT": 10,
    "NOV": 11,
    "DEZ": 12,
}


def load_corrections_factor_from_dict(factors, paid_month, paid_year, start_date=None):
    for month, year in factors:
        ref_date = datetime.date(year, month, 1).date()
        if not start_date or ref_date >= start_date:
            print(f"{month:02d}/{year:04d}: {factors[(month, year)]}", end="")
            cf, created = CorrectionFactor.objects.update_or_create(
                identifier="JEBRN",
                ref_payment_year=paid_year,
                ref_payment_month=paid_month,
                ref_difference_year=year,
                ref_difference_month=month,
                defaults={"factor": round(factors[(month, year)], 7)},
            )
            print(" OK" if created else f" \033[0;31mUPDATED\033[0m => {cf.factor}")


def load_corrections_factor_from_xls(path, paid_month, paid_year, start_date=None):

    wb = load_workbook(filename=path)
    sheet = wb.active
    factors = {}

    for row in sheet.rows:
        if row[0].value and row[0].value.date() >= start_date:
            ref_date = row[0].value.date()
            value = row[1].value
            print(f"{ref_date.month:02d}/{ref_date.year:04d}: {value}", end="")
            cf, created = CorrectionFactor.objects.update_or_create(
                identifier="JEBRN",
                ref_payment_year=paid_year,
                ref_payment_month=paid_month,
                ref_difference_year=ref_date.year,
                ref_difference_month=ref_date.month,
                defaults={"factor": value},
            )
            print(" U" if not created and "factor" in cf.diff else " C")
            factors[(ref_date.month, ref_date.year)] = value

    return factors


def load_corrections_factor_from_xls_pdf(path, paid_month, paid_year, start_date=None):

    wb = load_workbook(filename=path)
    sheet = wb["TABELA"]
    idx_years = {
        1: None,
        2: None,
        3: None,
        4: None,
        5: None,
        6: None,
        7: None,
        8: None,
        9: None,
        10: None,
    }
    factors = {}

    for row in sheet.rows:
        if not row[0].value:
            for r in row:
                if r.value:
                    idx_years[r.col_idx - 1] = r.value
        else:
            for r in row:
                year = idx_years.get(r.col_idx - 1)
                month = idx_months.get(row[0].value, 0)
                if r.value and year:
                    ref_date = datetime.date(year, month, 1)
                    if not start_date or ref_date >= start_date:
                        print(f"{month:02d}/{year:04d}: {r.value}", end="")
                        cf = CorrectionFactor.objects.filter(
                            identifier="JEBRN",
                            ref_payment_year=paid_year,
                            ref_payment_month=paid_month,
                            ref_difference_year=year,
                            ref_difference_month=month,
                        ).exclude(factor=r.value)
                        factors[(month, year)] = r.value
                        print(" ERRO" if cf else " OK")

    return factors


def load_correction_factor(path, start_date=None):

    with codecs.open(path, "r") as csv_file:
        lines = csv_file.readlines()

    conflicts = []
    for line in lines:
        str_date, str_factor = line.split(";")
        fdate = datetime.datetime.strptime(str_date, "%d/%m/%Y")
        factor = float(str_factor.replace("\n", "").replace(",", "."))
        cf = CorrectionFactor.objects.filter(
            identifier="JEBRN",
            ref_payment_year=2022,
            ref_payment_month=6,
            ref_difference_year=fdate.year,
            ref_difference_month=fdate.month,
        ).exclude(factor=factor)
    if cf:
        print(fdate.strftime("%m/%Y"), factor, cf.first())
