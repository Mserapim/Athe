import csv
import importlib
import xlwt
import os

# import pdfrw
from datetime import date
import pdfkit
import traceback
import pandas as pd
import io

from contrib.utils import getLogger, get_json_engine
from contrib.middleware import get_current_user
from default.websocket import RemoteEmmiter
from ged.models import Arquivo
from rh.queryregistration.ged_file import GedFile
from pathlib import Path
import hashlib
from django.conf import settings

# from odf.opendocument import load
# from odf import text, teletype
from django.template.loader import render_to_string
from io import BytesIO
import pypandoc
import tempfile
from docx import Document
from docx.shared import RGBColor
from django.apps import apps
from django.utils.module_loading import import_string
from apiv2.baseserializers import get_serializer_dinamico
from apiv2.utils import get_titulo_campo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pdf2docx import Converter
import subprocess

log = getLogger(__name__)
json = get_json_engine()

STYLE_HEAD_ROW = """
    align:
    wrap off,
    vert center,
    horiz center;
    borders:
    left THIN,
    right THIN,
    top THIN,
    bottom THIN;
    font:
    name Arial,
    bold on,
    colour_index gray80,
    height 0xA0;
    pattern:
    pattern solid,
    fore-colour 0x16;
"""

STYLE_DATA_ROW = """
    align:
    wrap on,
    vert center,
    horiz left;
    font:
    name Arial,
    bold off,
    height 0XA0;
    borders:
    left THIN,
    right THIN,
    top THIN,
    bottom THIN;
"""


def pdf_header_footer_options():
    options = {
        "--footer-right": "Hora: [time] | Página: [page] de [topage]",
        "--footer-font-size": "10",
        "--header-font-size": "10",
        "--footer-spacing": "12",
        "--margin-top": "5mm",
        "--margin-bottom": "25mm",
        "--margin-left": "5mm",
        "--margin-right": "5mm",
        "--footer-line": "",
        # '--header-html':f'{settings.RESOURCE_BASE_URL}/athenas/GFPCommitmentReport/viewer_header/'
    }
    return options


def create_file_xlsx(values_data, keys, file_path):
    """
    Função para criar um arquivo xlsx a partir de um dicionário de dados
    :param values_data: Dicionário de dados
    :param keys: Chaves do dicionário
    :param file_path: Caminho para salvar o arquivo
    """
    if keys:
        df = pd.DataFrame(values_data)
        df.rename(columns=keys, inplace=True)
    else:
        df = pd.DataFrame(values_data)

    if not file_path:
        output = io.BytesIO()
    else:
        output = file_path

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"Relatório")
        workbook = writer.book
        worksheet = workbook[f"Relatório"]
        for col_index, col_name in enumerate(df.columns, start=1):
            max_length = (
                max(len(str(col_name)), df[col_name].astype(str).map(len).max()) + 2
            )
            worksheet.column_dimensions[get_column_letter(col_index)].width = max_length

    if not file_path:
        output.seek(0)

    return output


def create_file_xls(values_data, filepath, keys):
    work_book = xlwt.Workbook(encoding="utf-8")
    work_book.owner = get_current_user().username
    key_sheet = 1
    for values in list(values_data):
        work_sheet = work_book.add_sheet("planilha" + str(key_sheet))
        style_head_row = xlwt.easyxf(STYLE_HEAD_ROW)
        style_data_row = xlwt.easyxf(STYLE_DATA_ROW)
        style_date_row = xlwt.easyxf(STYLE_DATA_ROW)
        style_date_row.num_format_str = "dd/mm/yyyy"
        # style_green = xlwt.easyxf("pattern: fore-colour 0x11, pattern solid;")
        # style_red = xlwt.easyxf(" pattern: fore-colour 0x0A, pattern solid;")
        for i in range(len(keys)):
            work_sheet.write(0, i, str(keys[i]).upper(), style_head_row)

        row = 1
        for value in values.get("data"):
            column = 0
            for item in value:
                data = value[item]
                colwidth = (
                    256 * len(str(data)) if 256 * len(str(data)) < 65536 else 65535
                )
                if colwidth > work_sheet.col(column).width:
                    work_sheet.col(column).width = colwidth
                if isinstance(data, date):
                    work_sheet.write(row, column, data, style_date_row)
                    column = column + 1
                if isinstance(data, dict):
                    for data_item in data:
                        data_value = data[data_item]
                        work_sheet.write(row, column, data_value, style_date_row)
                if isinstance(data, list):
                    for data_item in data:
                        if isinstance(data_item, dict):
                            for i in data_item:
                                d_value = data_item[i]
                                work_sheet.write(row, column, d_value, style_date_row)
                                column = column + 1
                            row = row + 1
                        else:
                            data_value = data[data_item]
                            work_sheet.write(row, column, data_value, style_date_row)
                        column = column + 1

                else:
                    work_sheet.write(row, column, data, style_data_row)
                column = column + 1
            row = row + 1
        key_sheet = key_sheet + 1
    if not filepath:
        return work_book.get_biff_data()

    return work_book.save(filepath)


def tmp_dir():
    return os.path.join(settings.UPLOAD_STORE_DIR, "reports")


def create_csv(values_data, filepath, keys, filename):
    if not filepath:
        filepath = os.path.join(tmp_dir(), filename)
        if not os.path.exists(tmp_dir()):
            os.makedirs(tmp_dir())
    with open(filepath, "w") as file:
        writer = csv.DictWriter(file, delimiter=";", fieldnames=keys)
        writer.writeheader()
        for data_item in values_data:
            for value in data_item.get("data"):
                writer.writerow(value)
    return repr(file)


def get_filename(filename, identifier):
    file = False
    try:
        log.info(
            Arquivo.objects.filter(filename=filename, user=get_current_user()).count()
        )
        file = Arquivo.objects.filter(filename=filename, user=get_current_user()).last()
        file = GedFile(file.file if file else filename, identifier)
    except Exception as e:
        log.error(e)
        log.info("Arquivo não encontrado!")
    return file if file else None


def create_gedfile(filename, buffer, mimetype, identifier):
    ged = None
    signature = Arquivo.hash_buffer(buffer)
    gedfile = GedFile(signature, identifier)
    ged, created = Arquivo.objects.update_or_create(
        file=signature,
        defaults={
            "filename": filename,
            "user": get_current_user(),
            "file": signature,
            "mimetype": mimetype,
            "acesso": 3,
        },
    )
    gedfile.save_file(gedfile, signature, buffer)
    ged.save(ignore_cache=True)
    return gedfile


def remote_emmiter(download, task, name=None, filename=None):
    if download and task.state != "failed":
        name = name if name else "base"
        RemoteEmmiter.emmit_for_user(
            task.owner,
            "base-report",
            path=f"/athenas/MPMTReports/download_file/?uuid={task.uuid}",
            filename=filename if filename else "",
        )


def run_context_data_function(path, class_name, params):
    lib = importlib.import_module(path)
    class_to_run = getattr(lib, class_name)
    return class_to_run.get_context_data(params)


# def write_pdf(input_pdf, data_dict, identifier):
#     ANNOT_KEY = '/Annots'
#     ANNOT_FIELD_KEY = '/T'
#     SUBTYPE_KEY = '/Subtype'
#     WIDGET_SUBTYPE_KEY = '/Widget'
#     template_pdf = pdfrw.PdfReader(os.path.join(Path.home(), input_pdf))
#     for page in template_pdf.pages:
#         for annotation in page[ANNOT_KEY]:
#             if annotation[SUBTYPE_KEY] == WIDGET_SUBTYPE_KEY:
#                 if annotation[ANNOT_FIELD_KEY]:
#                     key = annotation[ANNOT_FIELD_KEY][1:-1]
#                     if key in data_dict.keys():
#                         annotation.update(pdfrw.PdfDict(AP=''))
#                         # annotation.update(pdfrw.PdfDict(V='{}'.format(data_dict[key]),Ff=1))
#                         annotation.update(pdfrw.PdfDict(V='{}'.format(data_dict[key])))
#     template_pdf.Root.AcroForm.update(
#         pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true'))
#     )
#     directory_path = os.path.join(getattr(settings, 'UPLOAD_STORE_DIR', ''), identifier, get_current_user().username)
#     if not os.path.exists(directory_path):
#         os.makedirs(directory_path)
#     filepath = os.path.join(directory_path, hashlib.md5(str(template_pdf).encode('utf-8')).hexdigest())

#     pdfrw.PdfWriter().write(os.path.join(Path.home(), filepath), template_pdf)
#     return filepath


def write_odt(input_odt, data_dict, identifier):
    pass
    # odfpy não será mais utilizada, caso precise dessa função futuramente analisar outra lib
    # textdoc = load(input_odt)

    # for key in data_dict.keys():
    #     for item in textdoc.getElementsByType(text.P):
    #         ext_text = teletype.extractText(item)
    #         if ext_text.find(f'{key}') != -1:
    #             rep_text = ext_text.replace(f'{key}', f'{data_dict[key]}')
    #             new_item = text.P()
    #             new_item.setAttribute("stylename", item.getAttribute("stylename"))
    #             new_item.addText(rep_text)
    #             item.parentNode.insertBefore(new_item, item)
    #             item.parentNode.removeChild(item)

    # directory_path = os.path.join(getattr(settings, 'UPLOAD_STORE_DIR', ''), identifier, get_current_user().username)
    # if not os.path.exists(directory_path):
    #     os.makedirs(directory_path)
    # filename = hashlib.md5(str(f'{textdoc}.odt').encode('utf-8')).hexdigest()
    # filepath = f'{os.path.join(directory_path, filename)}'

    # textdoc.save(filepath)

    # return filepath


def criar_doc_docx(dados, html_path):
    """
    Preenche um template HTML com os dados fornecidos e cria um documento Word (.docx)

    Parâmetros:
    - dados (dict): Um dicionário contendo os dados a serem preenchidos no template HTML.
    - html_path (str): O caminho para o arquivo HTML que serve como template.
    """
    conteuto_html = render_to_string(html_path, dados)
    formato_entrada = "html"
    formato_saida = "docx"
    arquivo_temp = tempfile.NamedTemporaryFile(
        delete=False, suffix=f".{formato_saida}", mode="wb"
    )

    try:
        pypandoc.convert_text(
            conteuto_html,
            formato_saida,
            format=formato_entrada,
            outputfile=arquivo_temp.name,
        )

        with open(arquivo_temp.name, "rb") as f:
            saida_bytes = f.read()
            saida_bytesio = BytesIO(saida_bytes)

        doc = Document(saida_bytesio)

        for paragrafo in doc.paragraphs:
            if paragrafo.style.name.startswith("Heading"):
                for run in paragrafo.runs:
                    run.font.color.rgb = RGBColor(0, 0, 0)

        doc_bytesio = BytesIO()
        doc.save(doc_bytesio)

        return doc_bytesio.getvalue()

    finally:
        arquivo_temp.close()
        os.unlink(arquivo_temp.name)


def get_data_model_dinamico(data_model, campos):
    """
    Retorna os dados dinamicamente da api base de listagem

    Args:
    - data_model (dict): Um dicionário contendo dados do serializer e modelo.
    - campos (lits): lista de campos que serão filtrados
    Returns:
    - dict - dicionario com chaves e dados do modelo
    """

    model = apps.get_model(data_model["model"])
    queryset = model.objects.filter(id__in=data_model["ids"])
    serializer_class = import_string(data_model["serializer_name"])
    serializer_dinamico = get_serializer_dinamico(serializer_class, campos)
    serializer_data = serializer_dinamico(queryset, many=True).data
    dados = [
        {
            get_titulo_campo(serializer_class, campo): item.get(campo) or ""
            for campo in campos
        }
        for item in serializer_data
    ]
    data = {"data": dados, "keys": list(dados[0].keys())}
    return data


def criar_pdf(**kwargs):

    options = pdf_header_footer_options()
    filename = kwargs.get("filename")
    html_path = kwargs.get("html_path")
    mimetype = kwargs.get("mimetype")
    identifier = kwargs.get("identifier")
    params = kwargs.get("params")
    path = kwargs.get("path")
    class_name = kwargs.get("class_name")

    try:
        context_data = run_context_data_function(path, class_name, params)
        file = get_filename(filename, identifier)
        file_path = file.absolute_path if file else False
        html = render_to_string(html_path, context_data)
        output = pdfkit.from_string(html, output_path=file_path, options=options)
        if not file:
            file = create_gedfile(filename, output, mimetype, identifier)

        return file

    except Exception as err:
        log.exception(err)
        has_exception = err
        message = err

        error_message = traceback.format_exc()
        log.info(error_message)
        print(error_message)


def pdf_para_docx_em_memoria(pdf_bytes):
    docx_buffer = BytesIO()
    cv = Converter(stream=pdf_bytes)
    cv.convert(docx_buffer, start=0, end=None)
    cv.close()
    docx_buffer.seek(0)
    return docx_buffer


def gerar_arquivo_docx(html_path, context_data):
    html = render_to_string(html_path, context_data)
    options = pdf_header_footer_options()
    output = pdfkit.from_string(html, False, options=options)
    pdf_buffer = io.BytesIO(output)
    pdf_buffer.seek(0)
    docx = pdf_para_docx_em_memoria(output)
    return docx.getvalue()


def montar_servidor_lotacao_xlsx(lotacoes, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Servidores por Lotação"

    alinhamento_central = Alignment(horizontal="center", vertical="center")

    cabecalhos = [
        ("Ministério Público do Estado de Mato Grosso", 8),
        ("Procuradoria Geral de Justiça", 8),
        ("Relatório de Servidores por Lotação", 1),
    ]

    linha = 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=16)
    celula = ws.cell(row=linha, column=1, value=cabecalhos[0][0])
    celula.font = Font(bold=True)
    celula.alignment = alinhamento_central
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=16)
    celula = ws.cell(row=linha, column=1, value=cabecalhos[1][0])
    celula.font = Font(bold=True)
    celula.alignment = alinhamento_central
    linha += 2

    for lotacao in lotacoes:
        titulo_lotacao = f"{lotacao.get('serv_resp_matricula')} - {lotacao.get('serv_resp')} / {lotacao.get('nome')}"
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=16)
        ws.cell(row=linha, column=1, value=titulo_lotacao).font = Font(bold=True)
        linha += 1

        ws.cell(row=linha, column=1, value="SERVIDORES").font = Font(bold=True)
        linha += 1

        for servidor in lotacao.get("servidores", []):
            ws.cell(row=linha, column=1, value=servidor)
            linha += 1

        linha += 1

    for idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    if file_path:
        return wb.save(file_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
